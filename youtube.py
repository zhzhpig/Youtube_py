import requests
from langdetect import detect
from textblob import TextBlob
from snownlp import SnowNLP
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import matplotlib.pyplot as plt
import emoji
from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer
import tkinter as tk
import tkinter.font as tkFont
from tkinter import filedialog
from tkinter import ttk
from PIL import Image, ImageTk
from functools import partial
import logging
import os
import json
import sys
import threading
from nltk.tokenize import word_tokenize
from cnsenti import Sentiment
from wordcloud import WordCloud, STOPWORDS,ImageColorGenerator
import numpy as np
import jieba.posseg as pseg
from collections import Counter
import PIL.Image as Image
import random
import nltk
import re
import concurrent.futures
import webbrowser
from datetime import datetime
import matplotlib.patches as patches
import time

def remove_emojis(text):
    # 使用正则表达式匹配emoji，但不包括中文字符
    emoji_pattern = re.compile("(?![\u4e00-\u9fff])"+"["
        "\U0001F600-\U0001F64F"  # emoticons
        "\U0001F300-\U0001F5FF"  # symbols & pictographs
        "\U0001F680-\U0001F6FF"  # transport & map symbols
        "\U0001F1E0-\U0001F1FF"  # flags (iOS)
        "\U00002702-\U000027B0"
        "\U000024C2-\U0001F251"
        "]", flags=re.UNICODE)

    # 替换emoji为''
    return emoji_pattern.sub(r'', text)

def grey_color_func(word, font_size, position, orientation, random_state=None,
                    **kwargs):
        return "hsl(0, 0%%, %d%%)" % random.randint(50, 100)

def draw_wordcloud(df,name):
    logger.info("开始绘制词云图")
    nrc_emotion_column = df['评论内容']
    zh_text = []
    en_text = []

    # 遍历nrc_emotion_column的每一行
    for index, value in nrc_emotion_column.dropna().items():
        text = remove_emojis(value)
        try:
            # 检测文本的语言
            language = detect(text)
            if language == 'zh-cn' or language == 'zh':
                zh_text.append(text)
            elif language == 'en':
                en_text.append(text)
        except:
            # 如果检测失败，可以跳过这条记录
            pass

    zh_text_str = ' '.join(zh_text)
    en_text_str = ' '.join(en_text)
    # 分词并词性标注
    zh_tokens = [(word, flag) for word, flag in pseg.cut(zh_text_str)]
    en_tokens = [word for word in nltk.pos_tag(word_tokenize(en_text_str))]
    # 合并结果
    all_tokens = zh_tokens + en_tokens


    # 按指定长度和词性提取词
    report_words = []
    min_length = int(entry6.get())  # 最小长度
    max_length = int(entry7.get())  # 最大长度
    desired_pos = get_selected_pos()
    try:
        for word, flag in all_tokens:
            if (min_length<=len(word)<=max_length) and (flag in desired_pos): #这里设置统计的字数
                report_words.append(word)
    except Exception as e:
        logger.error(f"分词失败，因为：{e}")
        error_log.append(f"分词失败，因为：{e}")
        return
    logger.info(f"report_words:{report_words}")
    # 统计高频词汇
    result = Counter(report_words).most_common(200) #词的个数
    # 设置停用词
    stopwords = set(STOPWORDS)
    stopwords.update(entry5.get().split(','))
    # 建立词汇字典
    content = dict(result)
    filtered_content = {}

    # 遍历 content 字典中的每一个键值对
    for word, freq in content.items():
        # 如果当前词不在停用词集合中，则将其添加到 filtered_content 中
        if word not in stopwords:
            filtered_content[word] = freq

    # 现在 filtered_content 就是去除了停用词的字典
    
    logger.info(f"词云图词典:{filtered_content}\n词云图词典大小：{len(filtered_content)}")
    
    if radio_var.get():
        logger.info("使用用户自定义词云图遮罩图片，请选择图片路径...")
        try:
            Image_file_path = filedialog.askopenfilename(
                title="选择词云图遮罩图片",
                filetypes=(("JPEG文件", "*.jpg;*.jpeg"), ("PNG文件", "*.png"), ("JPG文件", "*.jpg"), ("所有文件（不选图片就报错）", "*.*")),
                initialdir=default_ico_directory
            )
            # 验证文件是否为有效的图片格式
            _, file_extension = os.path.splitext(Image_file_path)
            valid_extensions = ('.jpg', '.jpeg', '.png')
            
            if file_extension.lower() not in valid_extensions:
                raise ValueError("选择的文件不是有效的图片格式")

        except Exception as e:
            logger.error(f"用户自定义词云图遮罩图片路径错误，因为：{e}")
            error_log.append(f"用户自定义词云图遮罩图片路径错误，因为：{e}")
            return
    else:
        Image_file_path = f'{default_ico_directory}/usa_map.png'
        if os.path.exists(Image_file_path):
            logger.info(f"使用默认词云图遮罩图片，图片路径为：{Image_file_path}")
        else:
            logger.error(f"词云图遮罩图片默认路径{Image_file_path}不存在，请自定义词云图遮罩图片路径！")
            error_log.append(f"词云图遮罩图片默认路径{Image_file_path}不存在，请自定义词云图遮罩图片路径！")
            return

    #设置png掩膜（yourfile.png根据实际路径进行替换）
    try:
        background = Image.open(Image_file_path).convert("RGBA")
        Imgwidth, Imgheight = background.size
        mask = np.array(background)
    except Exception as e:
        logger.error(f'读取背景图片失败，因为：{e}')
        error_log.append(f'读取背景图片失败，因为：{e}')
        return
    # 提取背景图片颜色
    img_colors = ImageColorGenerator(mask)

    # 设置字体样式路径
    font_path = r"C:\Windows\Fonts\SIMLI.TTF"
    
    # 设置字体大小
    max_font_size =200
    min_font_size =10
    
    # 生成词云
    wordcloud = WordCloud(
                        scale=1,                         #输出清晰度
                        font_path=font_path,             #输出路径
                        width=Imgwidth,                  #输出图片宽度
                        height=Imgheight,                #输出图片高度
                        background_color='white',        #图片背景颜色
                        stopwords=stopwords,             #停用词
                        mask=mask,                       #掩膜
                        relative_scaling=0.5,            #控制词频和字体大小之间的关系
                        max_words=len(filtered_content), #最大词数
                        max_font_size=max_font_size,     #最大字体大小
                        collocations=False,              #避免一些大的词占据过多空间
                        min_font_size=min_font_size)     #最小字体大小
    try:
        wordcloud.generate_from_frequencies(filtered_content)
    except Exception as e:
        logger.error(f'词云图生成失败，因为：{e}')
        error_log.append(f'词云图生成失败，因为：{e}')
        return
    try:
        wordcloud.recolor(color_func=img_colors)
    except Exception as e:
        logger.error(f'词云图颜色设置失败，因为：{e}')
        error_log.append(f'词云图颜色设置失败，因为：{e}')
        return
    # 使用 matplotlib 显示词云
    fig = plt.figure(figsize=(15, 10))  # 调整figure的尺寸
    plt.imshow(wordcloud, interpolation='bilinear')
    plt.axis('off')
    # 保存词云图
    try:
        os.makedirs(default_img_directory, exist_ok=True)
        wordcloud.to_file(f"{default_img_directory}/{name}_wordcloud.png")
        logger.info(f"词云图绘制完成，图片保存地址为{default_img_directory}/{name}_wordcloud.png")
        plt.show()
        plt.close(fig)
    except Exception as e:
        logger.error(f'保存词云图图片失败，因为：{e}')
        error_log.append(f'保存词云图图片失败，因为：{e}')
        return
    
def draw_pip_chart(df,name):
    logger.info(f"开始绘制饼图")
    nrc_emotion_column = df['NRC情绪分析']
    # 计算每种情绪的出现次数
    emotion_counts = nrc_emotion_column.value_counts()

    # 将情绪的英文名称转换为中文
    emotion_counts_translated = emotion_counts.rename(emotion_translation)

    # 准备数据
    labels = emotion_counts_translated.index
    sizes = emotion_counts_translated.values

    # 计算百分比
    percentages = (sizes / sizes.sum()) * 100
    # 保留两位小数
    percentages_rounded = np.round(percentages, 2)

    # 分离主要数据项和较小的数据项
    threshold = 5  # 设置一个阈值，小于该阈值的合并为"次要情绪"
    main_sizes = sizes[percentages_rounded >= threshold]
    main_labels = labels[percentages_rounded >= threshold]
    min_sizes = sizes[percentages_rounded < threshold]
    min_labels =labels[percentages_rounded < threshold]
    other_sum = sum(sizes[percentages_rounded < threshold])
    
    # 如果有较小的数据项，将它们合并为"次要情绪"
    if other_sum > 0:
        main_sizes = list(main_sizes) + [other_sum]
        main_labels = list(main_labels) + ['次要情绪']
    
    try:
        # 绘制饼状图
        fig, ax = plt.subplots(figsize=(12, 8))  # 增加图表尺寸
        plt.rcParams['font.sans-serif'] = ['SimHei']  # 用黑体显示中文
        plt.rcParams['axes.unicode_minus'] = False   # 正常显示负号
        
        # 绘制外层饼图
        outer_wedges, outer_texts, outer_autotexts = ax.pie(main_sizes, labels=main_labels, autopct='%1.1f%%',
                                                            startangle=140, labeldistance=1.1,
                                                            wedgeprops=dict(width=0.3), pctdistance=0.85)
        
        # 设置百分比文本的样式
        for autotext in outer_autotexts:
            autotext.set_color('black')
            autotext.set_fontsize(12)
        
        # 设置标签文本的样式
        for text in outer_texts:
            text.set_fontsize(10)
        
        # 绘制内层饼图
        inner_wedges, inner_texts,inner_autotexts = ax.pie(min_sizes, labels=min_labels,autopct='%1.1f%%',radius=0.6,wedgeprops=dict(width=0.3),labeldistance=0.3)

        # 设置百分比文本的样式
        for autotext in inner_autotexts:
            autotext.set_color('black')
            autotext.set_fontsize(12)
        
        # 设置标签文本的样式
        for text in inner_texts:
            text.set_fontsize(10)
    except Exception as e:
        logger.error(f'绘制饼图失败，因为：{e}')
        error_log.append(f'绘制饼图失败，因为：{e}')
        return
    
    try:
        # 添加图例
        # 外层饼图的图例
        legend_labels_outer = main_labels  # 直接使用 main_labels 列表
        # 使用 patches.Patch 创建图例项
        legend_handles_outer = [patches.Patch(color=wedge.get_facecolor(), label=label)
                                for wedge, label in zip(outer_wedges, legend_labels_outer)]

        # 创建外层图例
        outer_legend = ax.legend(handles=legend_handles_outer, title="主要情绪",
                                loc='upper right', bbox_to_anchor=(1, 1), fontsize=10,
                                frameon=True, fancybox=True, shadow=True, facecolor='white',
                                edgecolor='black', labelcolor='black', ncol=1)

        ax.add_artist(outer_legend)
        # 内层饼图的图例
        inner_legend_labels = min_labels
        # 使用 patches.Patch 创建图例项
        legend_handles_inner = [patches.Patch(color=wedge.get_facecolor(), label=label)
                                for wedge, label in zip(inner_wedges, inner_legend_labels)]

        # 创建内层图例
        inner_legend = ax.legend(handles=legend_handles_inner, title="次要情绪",
                                loc='lower right', bbox_to_anchor=(1, 0.05), fontsize=10,
                                frameon=True, fancybox=True, shadow=True, facecolor='white',
                                edgecolor='black', labelcolor='black', ncol=1)
        ax.add_artist(inner_legend)
    except Exception as e:
        logger.error(f'饼图图例绘制失败，因为：{e}')
        error_log.append(f'饼图图例绘制失败，因为：{e}')

    ax.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.
    ax.set_title('情绪分析结果分布', fontsize=18, color='black',pad=30)

    
    try:
        os.makedirs(default_img_directory, exist_ok=True)
        plt.savefig(f'{default_img_directory}/{name}_情感分析结果（饼状图）.png')
        logger.info(f"饼图绘制完成，图片保存地址为{default_img_directory}/{name}_情感分析结果（饼状图）.png")
        plt.show()
        plt.close(fig)
    except Exception as e:
        logger.error(f'保存饼状图图片失败，因为：{e}')
        error_log.append(f'保存饼状图图片失败，因为：{e}')
        return
    
def analyze_nrc_emotion(text):
    # 加载NRC情感词库
    nrc_path = 'NRC/NRC-Emotion-Lexicon/NRC-Emotion-Lexicon-Wordlevel-v0.92.txt'
    with open(nrc_path, 'r', encoding='utf-8') as f:
        nrc_lexicon = pd.read_csv(f, sep='\t', names=['word', 'emotion', 'association'], skiprows=45)
    nrc_lexicon.set_index(['word', 'emotion'], inplace=True)

    # 将NRC情感词库转换为更易查询的格式
    emotions = ['anger', 'anticipation', 'disgust', 'fear', 'joy', 'negative', 'positive', 'sadness', 'surprise', 'trust']
    nrc_lexicon = nrc_lexicon.unstack().reset_index()
    nrc_lexicon.columns = ['word'] + emotions
    nrc_lexicon.set_index('word', inplace=True)
    nrc_lexicon.fillna(0, inplace=True)
    logger.info("-----NRC以及cnsenti情绪分析开始-----\n")
    logger.info(f"评论内容: {text}")
    emojis = ''.join(c for c in text if c in emoji.UNICODE_EMOJI)
    max_emotion = 'unknown'
    try:
        language = detect(text)
    except:
        if emojis=='':
            language = 'en'
        else:
            # 分析文本的情感（包括表情符号）
            language = 'emojis'
            emotion = emojis_analysis(text)
            if emotion.find("积极") != -1:
                max_emotion = 'positive'
            elif emotion.find("消极") != -1:
                max_emotion = 'negative'
            else:
                max_emotion = 'netural'
        logger.info(f"emojis分析")
        logger.info(f"情绪类型: {max_emotion}")
        logger.info("-----NRC以及cnsenti情绪分析结束-----\n")
        return max_emotion
    if language == 'zh-cn' or language == 'zh':

        # 使用cnsenti进行情绪分析
        sentiment = Sentiment()
        emotion_scores = sentiment.sentiment_calculate(text)
        # 创建一个新的字典，只包含 'pos' 和 'neg' 的键值对，并将键名替换
        filtered_scores = {
            'positive' if k == 'pos' else 'negative': v
            for k, v in emotion_scores.items()
            if k in ('pos', 'neg')
        }
        # 检查 'Positive' 和 'Negative' 的得分是否都为0
        if all(v == 0 for v in filtered_scores.values()):
            max_emotion = 'netural'
        else:
            # 找出 'Positive' 和 'Negative' 中得分最高的情绪
            max_emotion = max(filtered_scores.items(), key=lambda x: x[1])[0]

        logger.info(f"中文情绪得分: {emotion_scores}")
        logger.info(f"情绪是否为0: {all(v == 0 for v in filtered_scores.values())}")
    else:
        # 英文分词
        tokens = word_tokenize(text.lower())
        emotion_scores = {emotion: 0 for emotion in emotions}
        
        # 计算情感得分
        for token in tokens:
            if token in nrc_lexicon.index:
                for emotion in emotions:
                    emotion_scores[emotion] += nrc_lexicon.loc[token][emotion]
        
        # 找出得分最高的情绪
        if all(v == 0 for v in emotion_scores.values()):
            max_emotion = 'netural'
        else:
            max_emotion = max(emotion_scores, key=emotion_scores.get)
        logger.info(f"英文情绪得分: {emotion_scores}")
        logger.info(f"情绪是否为0: {all(v == 0 for v in emotion_scores.values())}")
        
    logger.info(f"情绪类型: {max_emotion}")
    logger.info("-----NRC以及cnsenti情绪分析结束-----\n")
    return max_emotion

def emojis_analysis(text):
    sentiment = analyzer.polarity_scores(text)
    if sentiment['compound'] > 0:
        return 'emojis积极'
    elif sentiment['compound'] < 0:
        return 'emojis消极'
    else:
        return 'emojis中立'

def analyze_sentiment(text):
    # 提取文本中的所有表情符号
    emojis = ''.join(c for c in text if c in emoji.UNICODE_EMOJI)
    logger.info(f"emojis:{emojis}")
    try:
        language = detect(text)
    except:
        if emojis=='':
            language = 'en'
        else:
            # 分析文本的情感（包括表情符号）
            language = 'emojis'
            return emojis_analysis(text)
            
    logger.info(f"language:{language}")

    if language == 'zh-cn' or language == 'zh':
        s = SnowNLP(text)
        sentiment = s.sentiments
        if sentiment >= 0.6:
            return 'SnowNLP积极'
        elif sentiment <= 0.4:
            return 'SnowNLP消极'
        else:
            return 'SnowNLP中立'
    else:
        analysis = TextBlob(text)
        sentiment = analysis.sentiment.polarity
        if sentiment > 0:
            return 'TextBlob积极'
        elif sentiment < 0:
            return 'TextBlob消极'
        else:
            return 'TextBlob中立'
    
def get_comments(comments,videoId,page_token):
    MAX_RESULTS = int(entry2.get())
    ORDER = selected_value.get()
    # 你的API Key
    API_KEY = entry4.get()
    # 构造请求URL
    url = f"https://www.googleapis.com/youtube/v3/commentThreads"
    params = {
        'part': 'snippet,replies',  # 请求评论的摘要信息和回复
        'videoId': videoId,
        'maxResults': MAX_RESULTS,          # 每次请求最多返回100条评论
        'order': ORDER,       # 根据相关性排序评论
        'textFormat': 'plainText',  # 返回纯文本格式的评论
        'key': API_KEY,              # 你的API Key
        'pageToken': page_token
    }
    # 发送GET请求
    while not stop_event.is_set():
        if stop_event.is_set():
            logger.info("重新请求，当前请求停止！")
            break
        try:
            response = requests.get(url, params,timeout=5)
            # 检查响应状态码
            if response.status_code == 200:
                data = response.json()
                # 打印评论
                for item in data['items']:
                    if stop_event.is_set():
                        logger.info("重新请求，打印评论停止！")
                        break
                    commentPd = {}
                    comment = item['snippet']['topLevelComment']
                    author = comment['snippet']['authorDisplayName']
                    text = comment['snippet']['textDisplay']
                    senti = analyze_sentiment(text)
                    nrc_max_emotion = analyze_nrc_emotion(text)
                    commentPd["评论者"]=author
                    commentPd["评论内容"]=text
                    commentPd["评论时间"]=comment['snippet']['publishedAt']
                    commentPd["点赞数"]=comment['snippet']['likeCount']
                    commentPd["情感分析"]=senti
                    commentPd["NRC情绪分析"]=nrc_max_emotion
                    comments.append(commentPd)
                    logger.info(f'评论者: {author}')
                    logger.info(f'情感分析: {senti}')
                    logger.info('-------------------------')
            else:
                logger.warning(f"Request failed with status code {response.status_code}")
                logger.error(f"响应非200，响应值为：{response.text}")
            return comments, data.get('nextPageToken')
        except Exception as e:
            logger.error(f"请求失败！错误为：{e}")
            error_log.append(f"请求失败！错误为：{e}")
            return comments, None

def main():
    # 线程的主要工作
    pageCout = 1
    row_count = 0
    pageNum = int(entry3.get())
    VIDEO_ID = entry1.get().strip()
    # 文件名称
    db_name = VIDEO_ID
    root = tk.Tk()
    root.withdraw()
    # 使用 filedialog.askdirectory() 方法让用户选择目录
    directory = False

    if directory:
        excel_file = f"{directory}/{db_name}.xlsx"
    elif not directory:
        os.makedirs(default_file_directory, exist_ok=True)
        excel_file = f"{default_file_directory}/{db_name}.xlsx"
        logger.info(f"使用默认路径.{default_file_directory}/{db_name}.xlsx")
    page_token = None
    logger.info(f'第一次传入page_token：{page_token}')
    comments, next_page_token = get_comments([],VIDEO_ID, page_token)

    while next_page_token and pageCout < pageNum:
        if stop_event.is_set():
            logger.info(f"重新请求，第{pageCout+1}次请求停止！")
            break
        if pageCout == 1:
            logger.info(f'第{pageCout+1}次传入next_page_token：{next_page_token}')
            newcomments, next_page_token = get_comments(comments,VIDEO_ID, next_page_token)
        else:
            logger.info(f'第{pageCout+1}次传入next_page_token：{next_page_token}')
            newcomments, next_page_token = get_comments(newcomments,VIDEO_ID, next_page_token)
        
        new_comments_data = newcomments
        new_comments_df = pd.DataFrame(new_comments_data)
        new_comments_df['评论时间'] = pd.to_datetime(new_comments_df['评论时间'], format='%Y-%m-%dT%H:%M:%SZ').dt.tz_localize(None)
        row_count = len(new_comments_df)
        try:
            new_comments_df.to_excel(excel_file, index=True)
        except Exception as e:
            logger.error(f"表格创建失败！错误为：{e}")
            error_log.append(f"表格创建失败！错误为：{e}")
        format_excel(new_comments_df,excel_file)
        logger.info(f'已爬取第{pageCout}页数据,共爬取{row_count}条数据！')
        pageCout += 1
        if next_page_token is None:
            logger.info(f'没有下一页了！')
            break
        elif stop_event.is_set():
            logger.info(f'已停止爬取！')
            break
    if stop_event.is_set():
        logger.info(f"线程结束完成！")
        clear_log()
    else:
        logger.info(f"全部评论爬取完成！总评论数为{row_count}条！")
        if row_count>0:
            draw_pip_chart(new_comments_df,db_name)
            draw_wordcloud(new_comments_df,db_name)
        logger.info(f"程序执行完成！")
        export_logs_to_file(default_log_directory,"error")
        export_logs_to_file(default_log_directory,"info")
        run_button.config(state=tk.NORMAL)
        stop_button.config(state=tk.NORMAL)
    
def run_program(run_button,stop_button):
    global error_log
    error_log = []
    # 设置按钮为不可点击状态
    run_button.config(state=tk.DISABLED)
    stop_button.config(state=tk.DISABLED)
    global global_thread, should_stop
    if global_thread and global_thread.is_alive():
        # 设置停止事件，尝试停止线程
        logger.info("正在停止线程...")
        stop_event.set()
        global_thread.join(timeout=1)  # 设置超时时间
        should_stop = True
    else:
        # 清除停止事件，开始新线程
        clear_log()
        logger.info("开始新线程...")
        stop_event.clear()
        should_stop = False
        global_thread = threading.Thread(target=main)
        global_thread.start()
# 计算每列的最大宽度，可以根据需要调整乘数因子，默认是1.2
def calculate_max_widths(df, multiplier=1.6):
    # 计算每列的最大字符长度
    max_widths = [max((df[col].astype(str).str.len().max(), len(col))) for col in df.columns]
    # 考虑到数字和日期等可能影响列宽，适当增加宽度
    return [width * multiplier for width in max_widths]

def format_excel(df,excel_file):
    # 加载Excel工作簿
    wb = Workbook()
    ws = wb.active

    # 读取DataFrame到工作表
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    header_row = 1  # 第一行是表头

    # 创建一个Font对象，设置加粗
    bold_font = openpyxl.styles.Font(bold=True)

    # 创建一个Alignment对象，设置居中对齐
    center_alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal='center', vertical='center')

    # 根据DataFrame内容设置列宽
    max_widths = calculate_max_widths(df)
    for idx, width in enumerate(max_widths):
        col_letter = openpyxl.utils.get_column_letter(idx + 1)
        min_col = openpyxl.utils.column_index_from_string(col_letter)  # 将列字母转换为列索引
        max_col = min_col
        if col_letter == 'B':
            ws.column_dimensions[col_letter].width = 100
            # 遍历指定列的所有单元格，包括表头
            for row in ws.iter_rows(min_row=1, max_col=max_col, min_col=min_col, values_only=False):
                for cell in row:
                    cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
        else:
            ws.column_dimensions[col_letter].width = width+5
            # 遍历指定列的所有单元格，包括表头
            for row in ws.iter_rows(min_row=1, max_col=max_col, min_col=min_col, values_only=False):
                for cell in row:
                    cell.alignment = center_alignment

    # 遍历表头的每一列
    for col_idx in range(ws.max_column):
        # 获取表头单元格
        cell = ws.cell(row=header_row, column=col_idx + 1)
        
        # 设置字体和对齐
        cell.font = bold_font
        cell.alignment = center_alignment

    # 保存修改
    wb.save(excel_file)

def export_logs_to_file(log_path,log_type):
    # 获取当前日期和时间
    now = datetime.now()
    # 格式化日期和时间为 yyyy-MM-dd HH:mm:ss
    formatted_now = now.strftime("%Y-%m-%d %H:%M:%S")
    save_path = os.path.join(log_path, f'{log_type}_logs.txt')
    logger.info(f"导出{log_type}日志,日志存放路径为：{save_path}")
    try:
        os.makedirs(log_path, exist_ok=True)
    except Exception as e:
        logger.error(f'创建日志目录{log_path}失败，因为：{e}')
        error_log.append(f'创建日志目录{log_path}失败，因为：{e}')
    if log_type == 'info':
        # 从 Text 控件中获取所有内容
        log_content = log_text.get('1.0', tk.END)
    elif log_type == 'error':
        log_content = error_log
    # 写入文件
    try:
        with open(save_path, 'a', encoding='utf-8') as file:
            if len(log_content)>0:
                file.write(f'日志创建时间：{formatted_now}\n')
                if isinstance(log_content, str):
                    file.write(log_content)
                elif isinstance(log_content, list):
                    for item in log_content:
                        file.write(item + '\n')
                else:
                    raise ValueError("无效的日志内容格式，请检查日志内容格式是否正确！")
    except Exception as e:
        logger.error(f'保存日志{save_path}失败，因为：{e}')
        error_log.append(f'保存日志{save_path}失败，因为：{e}')
    

def on_exit():
    # 执行清理操作...
    # 异步执行耗时操作
    with concurrent.futures.ThreadPoolExecutor() as executor:
        logger.info(f"执行程序退出...")
        future = executor.submit(long_running_operation)
        # 不等待结果，立即退出
        plt.close('all')
        root.quit()
        root.destroy()
        sys.exit()

def long_running_operation():
    global global_thread
    stop_event.set()
    global_thread.join(timeout=1)
    os._exit(0)

def load_questions_from_file(file_path):
    os.makedirs(default_question_usual_directory, exist_ok=True)
    if os.path.exists(file_path):
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                data = json.load(file)
                return data
        except Exception as e:
            logger.error(f"加载常见问题{file_path}文本错误: {e}")
            error_log.append(f"加载常见问题{file_path}文本错误: {e}")
    else:
        try:
            # 如果文件不存在，则创建文件并写入默认数据
            with open(file_path, 'w', encoding='utf-8') as file:
                json.dump(default_question_usual, file, indent=4)
            logger.info(f"创建并写入默认数据到{file_path}")
        except Exception as e:
            logger.error(f"创建常见问题{file_path}文本错误: {e}")
            error_log.append(f"创建常见问题{file_path}文本错误: {e}")


if __name__ == '__main__':
    # 全局变量
    global_thread = None
    stop_event = threading.Event()
    should_stop = False
    default_img_directory = 'analysisImg'
    default_file_directory = 'YouTubeData'
    default_log_directory = 'log'
    default_ico_directory = 'images'
    default_question_usual_directory = 'question_usual'
    default_question_usual_txt = 'question_usual.txt'
    icon_path = os.path.join(default_ico_directory, 'image.ico')
    error_log = []
    default_word_pos = ['n','d','NN','NNS','v','a','JJR','JJS','r','j']
    # 版本信息
    app_version = 'v1.9.3'
    app_title = f'爬YouTube评论软件{app_version} | 论文版'
    app_copyright = '版权所有 © 2024-2050 zhzhpig Technologies Co., Ltd. All rights reserved.'
    app_contact = 'zhzhpig@163.com'
    app_content = "本软件是一个用于分析YouTube评论的Python软件，可以分析评论情感、评论内容、评论时间、评论者，生成评论情感分析报告，并支持导出为Excel表格。"
    app_uplog = {
        'v1.0':'根据API爬取评论，生成表格',
        'v1.1':'增加emotion、snownlp、textBlob情感分析',
        'v1.2':'增加情绪饼状图绘制功能',
        'v1.3':'增加评论内容词云图绘制功能',
        'v1.4':'修复videoid空格过滤以及词云图绘制问题',
        'v1.5':'增加info、error日志导出功能',
        'v1.6':'增加词云图词性筛选以及词云图字符长度筛选功能',
        'v1.7':'增加菜单导航功能，修复一些已知bug',
        'v1.8':'增加常见问题和图片双击事件',
        'v1.9':'增加窗口绑定esc删除事件',
        'v1.9.1':'优化饼状图绘制的图形界面',
        'v1.9.2':'自动获取修改日期',
        'v1.9.3':'增加内外层饼状图，按百分比分离低频率数据'
    }
    default_question_usual = [
        {
            "pic_url": "http_error.jpg",
            "error_msg": "HTTPSConnectionPool",
            "error_msg_zh": "网络连接有问题，检查vpn"
        },
        {
            "pic_url": "data_error.png",
            "error_msg": "cannot access local variable 'data' where it is not associated with a value",
            "error_msg_zh": "接口参数错误，根据上面日志的reason定位是哪个参数错误，然后修改，如图是reason是API_KEY_INVALID，则说明是API_KEY无效，检查你的APIkey"
        }
    ]
    # 获取当前文件的路径
    file_path = sys.argv[0]
    # 获取文件的最后修改时间的时间戳
    last_modified_time = os.path.getmtime(file_path)

    app_last_update = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(last_modified_time))
    app_author = 'zhzhpig'
    # 初始化情感分析器
    analyzer = SentimentIntensityAnalyzer()
    # 定义英文到中文的情绪映射
    emotion_translation = {
        'anger': '愤怒',
        'anticipation': '期待',
        'disgust': '厌恶',
        'fear': '恐惧',
        'joy': '喜悦',
        'negative': '负面',
        'positive': '正面',
        'sadness': '悲伤',
        'surprise': '惊讶',
        'trust': '信任',
        'netural': '中性',
        'unknown': '未知'
    }
    # 存储所有 IntVar 变量
    int_vars = {}
    # 定义词性的选项
    pos_options = {
        "中文词性": [
            ('n', '名词 (中文)'),
            ('d', '副词 (中文)'),
            ('v', '动词 (中文)'),
            ('a', '形容词 (中文)'),
            ('r', '代词 (中文)'),
            ('p', '介词 (中文)'),
            ('c', '连词 (中文)'),
            ('u', '助词 (中文)'),
            ('e', '叹词 (中文)'),
            ('o', '拟声词 (中文)'),
            ('q', '量词 (中文)'),
            ('m', '数词 (中文)'),
            ('f', '方位词 (中文)'),
            ('s', '处所词 (中文)'),
            ('i', '成语 (中文)'),
            ('j', '简称 (中文)'),
            ('l', '习用语 (中文)'),
            ('h', '前缀 (中文)'),
            ('k', '后缀 (中文)'),
            ('x', '非语素字 (中文)'),
            ('w', '标点符号 (中文)'),
        ],
        "英文词性": [
            ('NN', '名词 (单数)'),
            ('NNS', '名词 (复数)'),
            ('NNP', '专有名词 (单数)'),
            ('NNPS', '专有名词 (复数)'),
            ('VB', '动词 (基本形式)'),
            ('VBD', '动词 (过去式)'),
            ('VBG', '动词 (现在分词)'),
            ('VBN', '动词 (过去分词)'),
            ('VBP', '动词 (非第三人称单数)'),
            ('VBZ', '动词 (第三人称单数)'),
            ('JJ', '形容词'),
            ('JJR', '形容词 (比较级)'),
            ('JJS', '形容词 (最高级)'),
            ('RB', '副词'),
            ('RBR', '副词 (比较级)'),
            ('RBS', '副词 (最高级)'),
            ('IN', '介词或从属连词'),
            ('DT', '定冠词'),
            ('PRP', '人称代词'),
            ('PRP$', '物主代词'),
            ('CD', '卡迪纳尔数字'),
            ('MD', '模态动词'),
        ]
    }

    # 创建主窗口
    root = tk.Tk()
    root.title(app_title)
    #聚焦事件
    def on_focus_in(event,widget,default_text):
        """当获得焦点时，如果文本等于默认文本，则清空文本"""
        if widget.get() == default_text:
            widget.delete(0, "end")  # 清除所有文本

    def on_focus_out(event,widget,default_text):
        """当失去焦点时，如果文本为空，则恢复默认文本"""
        if not widget.get():
            widget.insert(0, default_text)  # 插入默认文本

    def esc_destory(event,window):
        window.destroy()
    # 创建 Checkbutton 控件
    def create_checkbuttons(parent):
        row = 0
        col = 0
        for i in range(5):  # 设置每一列的最小宽度
            parent.columnconfigure(i, minsize=100)  # 设置每一列的最小宽度为100像素

        for category, options in pos_options.items():
            tk.Label(parent, text=f"{category}:").grid(row=row, column=0, columnspan=4, sticky='w',padx=10,pady=10)
            row += 1
            col = 1
            for idx,(pos, description) in enumerate(options):
                # 为每个 Checkbutton 创建一个 IntVar
                var = tk.IntVar()
                int_vars[pos] = var
                tk.Checkbutton(parent, text=description, variable=var, onvalue=1, offvalue=0).grid(row=row, column=col, sticky='w',padx=20,pady=5)
                col += 1
                if col >= 5:
                    col = 1
                    row += 1
                if idx == len(options)-1:  # 如果是最后一个选项，则添加一个空行
                    row += 1
                    col = 0
        return row
    # 获取选中的词性
    def get_selected_pos():
        selected_pos_values = [pos for pos, var in int_vars.items() if var.get() == 1]
        if not selected_pos_values:  # 如果 selected_pos 为空
            logger.info(f"用户未选择词性，使用默认词性: {default_word_pos}")
            selected_pos_values = default_word_pos
        return selected_pos_values

    def destroy_popup(popup):
        popup.destroy()
        selected_value = get_selected_pos()
        logger.info(f"词云图词性为: {selected_value}")
    # 创建按钮来获取选中的词性
    def show_popup():
        popup = tk.Toplevel(root)
        popup.title("选择词云图词性")
        popup.geometry("900x600")  # 设置弹出框的大小
        popup.iconbitmap(icon_path)
        # 绑定 ESC 键事件
        popup.bind("<Escape>", lambda e, p=popup: destroy_popup(p))
        # 创建 Canvas 和 Scrollbar
        canvas = tk.Canvas(popup)
        scrollbar = tk.Scrollbar(popup, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        # 配置 Canvas
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # 将 Frame 包含在 Canvas 中
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        
        # 更新 Canvas 的滚动区域
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        
        # 创建 Checkbutton 控件
        create_checkbuttons(scrollable_frame)
        
        # 创建确定按钮
        ttk.Button(popup, text="确定", command=lambda: destroy_popup(popup)).pack(side="bottom", pady=10)



    # 创建按钮来打开弹窗
    ttk.Button(root, text="选择词云图词性", command=show_popup).place(x=310, y=330,width=100,height=30)
    # 验证函数
    def validate_entry(text, min_max, widget, default_text):
        try:
            if text.strip() != '': 
                value = int(text)
            else:
                value = int(default_text)
        except ValueError as e:
            logger.info(f"输入的值并非整型数字：{e}")
        try:
            if value > 0:
                if min_max == 'min':
                    if value >= int(entry7.get()):
                        raise ValueError(f"最小长度不能大于等于最大长度")
                elif min_max == 'max':
                    if value <= int(entry6.get()):
                        raise ValueError(f"最大长度不能小于等于最小长度")
            else:
                raise ValueError(f"长度必须大于0")
        except Exception as e:
            logger.info(f'输入错误：{e}')
        return True

    def do_open_url(url):
        webbrowser.open_new_tab(url)
    def show_help():
        about_window = tk.Toplevel(root)
        about_window.title("关于我们")
        about_window.geometry("500x400")  # 增加窗口大小
        about_window.resizable(False, False)
        about_window.iconbitmap(icon_path)
        # 绑定 ESC 键事件
        about_window.bind("<Escape>", lambda event, p=about_window:esc_destory(event,p))
        # 设置最大宽度
        max_width = 360  # 单元格的最大宽度

        # 创建一个框架来容纳所有元素，并设置背景颜色
        about_frame = ttk.Frame(about_window, padding=20, style="white.TFrame")  # 增加内边距
        about_frame.pack(fill=tk.BOTH, expand=True)

        # 加载应用图标
        icon_image = Image.open(icon_path)  # 替换为您的图标路径
        icon_image = icon_image.resize((50, 50), Image.Resampling.LANCZOS)
        app_icon = ImageTk.PhotoImage(icon_image)

        # 创建一个标签来显示应用图标
        icon_label = ttk.Label(about_frame, image=app_icon,style="white.TLabel")
        icon_label.image = app_icon  # 保持引用以避免垃圾回收
        icon_label.grid(row=0, column=0, sticky=tk.W, padx=(0, 20))  # 增加右侧外边距

        # 创建一个标签来显示应用信息
        about_label = ttk.Label(about_frame, text=app_title, font=("仿宋", 14, "bold"),style="white.TLabel")
        about_label.grid(row=0, column=1, sticky=tk.W, pady=(10, 0))

        # 创建一个标签来显示年份和更新时间
        version_label = ttk.Label(about_frame, text=f"{app_copyright} | 最后更新于 {app_last_update}", font=menu_font,style="white.TLabel", wraplength=max_width)
        version_label.grid(row=1, column=1, sticky=tk.W, pady=(10, 0))

        # 创建一个标签介绍功能
        website_label = ttk.Label(about_frame, text=f"软件功能介绍：{app_content}", cursor="hand2", foreground="blue", font=menu_font,style="white.TLabel", wraplength=max_width)
        website_label.grid(row=2, column=1, sticky=tk.W, pady=(10, 0))

        # 添加一条分割线
        separator = ttk.Separator(about_frame, orient=tk.HORIZONTAL)
        separator.grid(row=3, column=1, sticky=tk.EW, pady=(20, 0))

        # 创建一个标签来显示联系方式
        contact_label = ttk.Label(about_frame, text=f"联系我们：{app_contact}", font=menu_font,style="white.TLabel", wraplength=max_width)
        contact_label.grid(row=4, column=1, sticky=tk.W, pady=(10, 0))

        # 添加作者信息
        author_label = ttk.Label(about_frame, text=f"作者：{app_author}", font=menu_font,style="white.TLabel", wraplength=max_width)
        author_label.grid(row=5, column=1, sticky=tk.W, pady=(10, 0))
        # 创建确定按钮
        ttk.Button(about_window, text="关闭", command=lambda:about_window.destroy()).pack(side="bottom", pady=10)

    def show_update_log():
        update_log_window = tk.Toplevel(root)
        update_log_window.title("更新日志")
        update_log_window.geometry("500x500")
        update_log_window.resizable(False, False)
        update_log_window.iconbitmap(icon_path)
        # 绑定 ESC 键事件
        update_log_window.bind("<Escape>", lambda event,p=update_log_window:esc_destory(event,p))
        # 创建一个框架来容纳所有元素，并设置背景颜色
        update_log_frame = ttk.Frame(update_log_window, padding=20, style="white.TFrame")
        update_log_frame.pack(fill=tk.BOTH, expand=True)

        # 创建一个滚动条
        scrollbar = ttk.Scrollbar(update_log_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # 创建一个文本框来显示更新日志
        update_log_text = tk.Text(update_log_frame, wrap=tk.WORD, yscrollcommand=scrollbar.set, font=menu_font,border=None)
        update_log_text.pack(fill=tk.BOTH, expand=True)

        # 配置滚动条
        scrollbar.config(command=update_log_text.yview)

        # 将更新日志填充到文本框中
        for version, log in app_uplog.items():
            update_log_text.insert(tk.END, f"版本号：{version}\n更新内容：{log}\n\n")

        # 禁止编辑文本框
        update_log_text.config(state=tk.DISABLED)
        # 创建确定按钮
        ttk.Button(update_log_window, text="关闭", command=lambda:update_log_window.destroy()).pack(side="bottom", pady=10)

    def show_question_usual():
        question_usual = load_questions_from_file(os.path.join(default_question_usual_directory, default_question_usual_txt)) or default_question_usual
        question_window = tk.Toplevel(root)
        question_window.title("常见问题")
        question_window.geometry("800x500")
        question_window.resizable(False, False)
        question_window.iconbitmap(icon_path)
        max_width = 500
        # 绑定 ESC 键事件
        question_window.bind("<Escape>", lambda e,p=question_window:esc_destory(e,p))
        # 创建一个框架来容纳所有元素，并设置背景颜色
        question_usual_frame = ttk.Frame(question_window, padding=20, style="white.TFrame")
        question_usual_frame.pack(fill=tk.BOTH, expand=True)

        # 绑定双击事件
        def show_full_size_image(event,pic_url):
            # 加载原始大小的图片
            original_img = Image.open(pic_url)
            # 获取图片的宽度
            width, height = original_img.size
            # 设置新窗口的宽度与图片宽度相同
            new_window_width = width + 20  # 添加一些额外的空间
            new_window_height = height  # 可以根据需要调整高度
            
            # 打开一个新窗口显示全尺寸图片
            full_size_image_window = tk.Toplevel(question_window)
            full_size_image_window.title("查看图片")
            full_size_image_window.geometry(f"{new_window_width}x{new_window_height}")
            full_size_image_window.iconbitmap(icon_path)
            full_size_image_window.resizable(False, False)

            
            original_photo = ImageTk.PhotoImage(original_img)

            # 显示图片
            tk.Label(full_size_image_window, image=original_photo).pack()
            full_size_image_window.mainloop()

         # 创建一个滚动条
        scrollbar = ttk.Scrollbar(question_usual_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # 创建一个文本框来显示
        question_usual_text = tk.Text(question_usual_frame, wrap=tk.WORD, yscrollcommand=scrollbar.set, font=menu_font,border=None)
        question_usual_text.pack(fill=tk.BOTH, expand=True)
        # 配置滚动条
        scrollbar.config(command=question_usual_text.yview)

        for index in range(len(question_usual)):
            error_info = question_usual[index]
            
            error_msg = error_info["error_msg"]
            error_msg_zh = error_info["error_msg_zh"]

            # 创建一个 Frame 来容纳图片和文字
            frame = tk.Frame(question_usual_text)
            frame.pack(fill=tk.X, padx=5, pady=5)
            try:
                if error_info["pic_url"]:
                    pic_url = os.path.join(default_question_usual_directory,error_info["pic_url"])
                    # 加载图片
                    img = Image.open(pic_url)
                    img = img.resize((200, 150))  # 调整图片大小
                    photo = ImageTk.PhotoImage(img)
                    # 创建 Label 来显示图片
                    label_img = tk.Label(frame, image=photo)
                    label_img.image = photo  # 保存引用
                    label_img.pack(side=tk.LEFT, padx=5, pady=5)
                    # 绑定双击事件到图片标签
                    label_img.bind("<Double-Button-1>", lambda event, p=pic_url: show_full_size_image(event, p))
            except Exception as e:
                logger.error(f"获取常见问题的图片路径错误: {e}")
                error_log.append(f"获取常见问题的图片路径错误: {e}")

            # 创建 Label 来显示文字
            label_text = tk.Label(frame, text=f"错误信息：{error_msg}\n解决建议：{error_msg_zh}", justify=tk.LEFT,wraplength=max_width)
            label_text.pack(side=tk.LEFT, padx=5, pady=5)

            # 在 Text 组件中插入 Frame
            question_usual_text.window_create(tk.END, window=frame)
            question_usual_text.insert(tk.END, "\n\n")



        # 禁止编辑文本框
        question_usual_text.config(state=tk.DISABLED)
        # 创建确定按钮
        ttk.Button(question_window, text="关闭", command=lambda:question_window.destroy()).pack(side="bottom", pady=10)


    # 设置样式
    style = ttk.Style()
    style.configure("white.TFrame",background="white")  # 设置框架的背景颜色
    style.configure("white.TLabel",background="white")

    # 设置菜单栏的字体
    menu_font = tkFont.Font(family="仿宋", size=12)

    # 创建一个菜单栏
    menu_bar = tk.Menu(root,tearoff=0,font=menu_font)

    # 工具下拉菜单
    tool_menu = tk.Menu(menu_bar, tearoff=0)
    tool_menu.add_command(label="API Key获取教程",command=lambda:do_open_url('https://www.bilibili.com/read/cv34316980/'))
    tool_menu.add_command(label="下载VPN v2rayN获取教程",command=lambda:do_open_url('https://pan.baidu.com/s/12uatmjjGbfLPogUMzw3g8Q?pwd=zhzb'))
    tool_menu.add_separator()  # 添加一条分割线
    tool_menu.add_command(label="关闭", command=root.quit,accelerator='CTRL+Q')

    # 帮助菜单
    help_menu = tk.Menu(menu_bar, tearoff=0)
    help_menu.add_command(label="关于我们",command=show_help,accelerator='CTRL+B')
    help_menu.add_cascade(label="常见问题",command=show_question_usual,accelerator='CTRL+H')
    help_menu.add_command(label="更新日志",command=show_update_log,accelerator='CTRL+U')


    # 将下拉菜单添加到菜单栏
    menu_bar.add_cascade(label="工具", menu=tool_menu)
    menu_bar.add_cascade(label="帮助", menu=help_menu)
    # 配置主窗口使用这个菜单栏
    root.config(menu=menu_bar)
    # 正确地处理事件参数
    def on_ctrl_b(event):
        show_help()

    def on_ctrl_u(event):
        show_update_log()

    def on_ctrl_q(event):
        on_exit()

    def on_ctrl_h(event):
        show_question_usual()
        
    root.bind("<Control-b>", on_ctrl_b)
    root.bind("<Control-u>", on_ctrl_u)
    root.bind("<Control-q>", on_ctrl_q)
    root.bind("<Control-h>", on_ctrl_h)
    
    # 版权信息
    copyright = tk.Label(root, text=app_copyright, font=('仿宋', 10), fg='grey')
    copyright.place(x=180, y=685)
    # 设置窗口大小
    root.minsize(width=850, height=710)
    root.resizable(False, False)
    # 获取屏幕尺寸
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    # 左上角图标
    root.iconbitmap(icon_path)

    # 创建第一个输入框
    entry4_label = tk.Label(root,justify='left',text="APIkey:").place(x=30, y=30,width=100,height=30)
    tk.Label(root, justify='left', text='谷歌YouTube的APIKey', fg='red', ).place(x=480, y=30) 
    entry4 = tk.Entry(root)
    entry4.insert(0,'AIzaSyAqexAFKpWPgDR-4T_Fip2TWRBu9yIww9c')
    entry4.place(x=150, y=30,width=310,height=30)
    # 创建第二个输入框
    entry1_label = tk.Label(root,justify='left',text="YouTube视频ID:").place(x=30, y=90,width=100,height=30)
    entry1 = tk.Entry(root)
    entry1.insert(0,'ecxUsur_YLM')
    entry1.place(x=150, y=90,width=150,height=30)

    # 创建第三个输入框
    entry2_label = tk.Label(root,justify='left',text="每页的评论数:").place(x=320, y=90,width=100,height=30)
    tk.Label(root, justify='left', text='包括1到100（含 0 和 100）', fg='red', ).place(x=610, y=90) 
    entry2 = tk.Entry(root)
    entry2.insert(0,'20')
    entry2.place(x=440, y=90,width=150,height=30)

    # 创建第四个输入框
    entry3_label = tk.Label(root,justify='left',text="请求总页数:").place(x=30, y=150,width=100,height=30)
    tk.Label(root, justify='left', text='别请求太多，会崩，最多10页', fg='red', ).place(x=320, y=150) 
    entry3 = tk.Entry(root)
    entry3.insert(0,'2')
    entry3.place(x=150, y=150,width=150,height=30)

    # 创建 Combobox 控件
    selected_value = tk.StringVar()
    options = ['time', 'relevance']
    menu_text = tk.Label(root, text="评论排序方式:").place(x=30,y=210,width=100,height=30)
    combobox = ttk.Combobox(root, textvariable=selected_value, values=options, state='readonly')
    combobox.set(options[0])  # 设置默认值
    combobox.place(x=150, y=210, width=150, height=30)
    style = ttk.Style()
    style.configure('Custom.TCombobox', background='white')
    tk.Label(root, justify='left', text='time-按时间排序; relevance-按相关性排序', fg='red', ).place(x=320, y=210) 

    # 创建第五个输入框
    entry5_label = tk.Label(root,justify='left',text="词云图禁用词:").place(x=30, y=270,width=100,height=30)
    tk.Label(root, justify='left', text="多个词用英文逗号分隔',' 英文！英文！英文！！", fg='red', ).place(x=480, y=270) 
    entry5 = tk.Entry(root)
    entry5_default_text = '比如,example'
    entry5.insert(0,entry5_default_text)
    # 绑定事件处理器
    entry5.bind("<FocusIn>", partial(on_focus_in,widget=entry5, default_text=entry5_default_text))
    entry5.bind("<FocusOut>", partial(on_focus_out,widget=entry5, default_text=entry5_default_text))
    entry5.place(x=150, y=270,width=310,height=30)

    #创建单选按钮判断词云图是否自选背景
    radio_var = tk.IntVar()
    radio_var.set(False)
    tk.Label(root, justify='left', text='词云图背景是否自选:').place(x=30, y=330)
    rb1 = tk.Radiobutton(root, text="是", variable=radio_var, value=True)
    rb2 = tk.Radiobutton(root, text="否", variable=radio_var, value=False)
    rb1.place(x=150, y=330,width=50,height=30)
    rb2.place(x=230, y=330,width=50,height=30)

    
    # 创建第六个输入框
    entry6_label = tk.Label(root, justify='left', text="字符最小长度:")
    entry6_label.place(x=440, y=330, width=100, height=30)
    entry6 = tk.Entry(root, validate="key")
    entry6_default_text = '2'
    entry6.insert(0, entry6_default_text)
    entry6['validatecommand'] = (entry6.register(partial(validate_entry, min_max='min',widget=entry6, default_text=entry6_default_text)), '%P')
    entry6.place(x=540, y=330, width=30, height=30)


    
    # 创建第七个输入框
    entry7_label = tk.Label(root, justify='left', text="字符最大长度:")
    entry7_label.place(x=590, y=330, width=100, height=30)
    entry7 = tk.Entry(root, validate="key")
    entry7_default_text = '7'
    entry7.insert(0, entry7_default_text)
    entry7['validatecommand'] = (entry7.register(partial(validate_entry, min_max='max',widget=entry7, default_text=entry7_default_text)), '%P')
    entry7.place(x=690, y=330, width=30, height=30)

    # 创建一个 Frame
    tk.Label(root, justify='left', text='运行日志:').place(x=30, y=380)
    frame = tk.Frame(root)
    frame.place(x=30,y=420, width=780, height=200)
    # 在 Frame 内创建一个 Text 控件
    log_text = tk.Text(frame, wrap=tk.WORD, height=10, width=50)
    log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    # 创建一个滚动条
    scrollbar = tk.Scrollbar(frame, command=log_text.yview)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    # 配置 Text 控件使用滚动条
    log_text.config(yscrollcommand=scrollbar.set)
    # 禁止用户编辑 Text 控件
    log_text.config(state=tk.DISABLED)
    # 创建一个日志处理器，用于将日志信息写入 Text 控件
    class TextHandler(logging.Handler):
        def __init__(self, text_widget):
            super().__init__()
            self.text_widget = text_widget

        def emit(self, record):
            msg = self.format(record)
            self.text_widget.config(state=tk.NORMAL)
            self.text_widget.insert(tk.END, msg + '\n')
            self.text_widget.config(state=tk.DISABLED)
            # 自动滚动到底部
            self.text_widget.yview(tk.END)
            # 强制GUI更新
            self.text_widget.update_idletasks()

    # 配置日志
    logger = logging.getLogger(__name__)
    handler = TextHandler(log_text)
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    handler.setFormatter(formatter)
    handler.flush = lambda: sys.stdout.flush()  # 强制刷新stdout
    logger.addHandler(handler)
    logger.setLevel(logging.INFO)

    def clear_log():
        log_text.config(state=tk.NORMAL)
        log_text.delete(1.0, tk.END)
        log_text.config(state=tk.DISABLED)

    # 创建按钮，点击后运行程序
    run_button = tk.Button(root, text="开始请求")
    # 创建按钮，点击后结束程序
    stop_button = tk.Button(root, text="结束程序")
    run_button.config(command=lambda: run_program(run_button,stop_button))
    run_button.place(x=245, y=640, width=150, height=30)
    stop_button.config(command=on_exit)
    stop_button.place(x=435, y=640, width=150, height=30)
    
    root.protocol("WM_DELETE_WINDOW", on_exit)
    # 运行主循环
    root.mainloop()